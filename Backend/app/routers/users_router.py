from fastapi import APIRouter, HTTPException, Query
from typing import Optional, List
from pydantic import BaseModel, EmailStr
from datetime import datetime
from app.database import get_db_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from io import BytesIO
import hashlib
import binascii
import secrets

router = APIRouter(prefix="/api/users", tags=["users"])

# Pydantic models for request/response
class UserCreate(BaseModel):
    emp_id: int
    name: str
    email: EmailStr
    department: Optional[str] = None
    designation: Optional[str] = None
    hod: Optional[str] = None
    supervisor: Optional[str] = None
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    department: Optional[str] = None
    designation: Optional[str] = None
    hod: Optional[str] = None
    supervisor: Optional[str] = None

class UserResponse(BaseModel):
    id: int
    emp_id: int
    name: str
    email: str
    department: Optional[str]
    designation: Optional[str]
    hod: Optional[str]
    supervisor: Optional[str]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

def hash_password(password: str, salt: Optional[str] = None) -> tuple:
    """
    PBKDF2-HMAC-SHA256 hashing to match auth_router.py.
    Returns (password_hash, salt).
    """
    if salt is None:
        salt = binascii.hexlify(os.urandom(16)).decode()
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
    return binascii.hexlify(dk).decode(), salt

@router.get("/", response_model=List[UserResponse])
def get_all_users():
    """Get all users"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, emp_id, name, email, department, designation, hod, supervisor, created_at, updated_at
                    FROM users
                    ORDER BY emp_id
                """)
                users = cursor.fetchall()
        finally:
            connection.close()

        if not users:
            return []

        return users
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export-to-excel")
def export_users_to_excel():
    """Export all users to Excel file"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, emp_id, name, email, department, designation, hod, supervisor, created_at, updated_at
                    FROM users
                    ORDER BY emp_id
                """)
                users = cursor.fetchall()
        finally:
            connection.close()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        # Header row with styling
        headers = ["ID", "Employee ID", "Name", "Email", "Department", "Designation", "HOD", "Supervisor", "Created At", "Updated At"]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add data rows
        for user in users:
            ws.append([
                user.get('id'),
                user.get('emp_id'),
                user.get('name'),
                user.get('email'),
                user.get('department', ''),
                user.get('designation', ''),
                user.get('hod', ''),
                user.get('supervisor', ''),
                user.get('created_at'),
                user.get('updated_at')
            ])

        # Adjust column widths
        column_widths = [8, 15, 20, 25, 20, 20, 15, 20, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Center align all data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file content (binary) and filename in JSON for your frontend to consume
        return {
            "success": True,
            "message": "Users exported to Excel successfully",
            "file": output.getvalue(),
            "filename": f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{emp_id}", response_model=UserResponse)
def get_user_by_emp_id(emp_id: int):
    """Get user by employee ID"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT id, emp_id, name, email, department, designation, hod, supervisor, created_at, updated_at
                    FROM users
                    WHERE emp_id = %s
                """, (emp_id,))
                user = cursor.fetchone()
        finally:
            connection.close()

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        return user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{emp_id}", response_model=UserResponse)
def update_user(emp_id: int, user_data: UserUpdate):
    """Update user by employee ID"""
    try:
        # Check if user exists
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id FROM users WHERE emp_id = %s", (emp_id,))
                existing_user = cursor.fetchone()
        finally:
            connection.close()

        if not existing_user:
            raise HTTPException(status_code=404, detail="User not found")

        # Build update query dynamically
        update_fields = []
        update_values = []

        if user_data.name is not None:
            update_fields.append("name = %s")
            update_values.append(user_data.name)
        if user_data.email is not None:
            update_fields.append("email = %s")
            update_values.append(user_data.email)
        if user_data.department is not None:
            update_fields.append("department = %s")
            update_values.append(user_data.department)
        if user_data.designation is not None:
            update_fields.append("designation = %s")
            update_values.append(user_data.designation)
        if user_data.hod is not None:
            update_fields.append("hod = %s")
            update_values.append(user_data.hod)
        if user_data.supervisor is not None:
            update_fields.append("supervisor = %s")
            update_values.append(user_data.supervisor)

        if not update_fields:
            raise HTTPException(status_code=400, detail="No fields to update")

        # Always update updated_at
        update_fields.append("updated_at = CURRENT_TIMESTAMP")
        update_values.append(emp_id)

        # Execute update
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                sql = f"UPDATE users SET {', '.join(update_fields)} WHERE emp_id = %s"
                cursor.execute(sql, update_values)
                connection.commit()

                # Fetch updated user
                cursor.execute("""
                    SELECT id, emp_id, name, email, department, designation, hod, supervisor, created_at, updated_at
                    FROM users
                    WHERE emp_id = %s
                """, (emp_id,))
                updated_user = cursor.fetchone()
        finally:
            connection.close()

        return updated_user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{emp_id}")
def delete_user(emp_id: int):
    """Delete user by employee ID"""
    try:
        # Check if user exists
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id, name FROM users WHERE emp_id = %s", (emp_id,))
                user = cursor.fetchone()
        finally:
            connection.close()

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        # Delete user
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("DELETE FROM users WHERE emp_id = %s", (emp_id,))
                connection.commit()

                if cursor.rowcount == 0:
                    raise HTTPException(status_code=400, detail="Failed to delete user")
        finally:
            connection.close()

        return {
            "success": True,
            "message": f"User '{user['name']}' (emp_id: {emp_id}) deleted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Optional: endpoint to create user via this router (keeps hashing consistent).
# Remove if you don't want this route.
@router.post("/create")
def create_user(payload: UserCreate):
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id FROM users WHERE email = %s", (payload.email,))
                if cursor.fetchone():
                    raise HTTPException(status_code=409, detail="User already exists")

                pwd_hash, salt = hash_password(payload.password)
                cursor.execute("""
                    INSERT INTO users (emp_id, name, email, department, designation, hod, supervisor, password_hash, password_salt)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (payload.emp_id, payload.name, payload.email, payload.department, payload.designation, payload.hod, payload.supervisor, pwd_hash, salt))
                connection.commit()
                user_id = cursor.lastrowid
                cursor.execute("""
                    SELECT id, emp_id, name, email, department, designation, hod, supervisor, created_at, updated_at
                    FROM users WHERE id = %s
                """, (user_id,))
                user = cursor.fetchone()
        finally:
            connection.close()
        return {"success": True, "message": "User created", "data": user}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
